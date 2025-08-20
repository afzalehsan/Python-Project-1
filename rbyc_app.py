from tkinter import *
from tkinter import messagebox
import webbrowser as wb
import openpyxl
import os

window = Tk()
# Window Title
window.title("Member Registration Portal")
window.geometry("600x400")
# window. resizable (False, False)
# Functions commands
def register():
    name=name_info.get()
    age=age_info.get()
    email=email_info.get()
    phone=phone_info.get()
    if name=="":
        messagebox.showinfo("Information!", "Please Enter Your Name") 
    elif age=="":
        messagebox.showinfo("Information!", "Please Enter Your Age") 
    elif email=="":
        messagebox.showinfo("Information!", "please Enter Your E-mail")
    elif phone=="":
        messagebox.showinfo("Information!", "Please Enter Your Phone Number")
    else:
        try:
            file_path = "rbyc_members_info.xlsx"

            # Load Excel file
            if os.path.exists(file_path):
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active

                # Find the next available row
                next_row = sheet.max_row + 1
                serial_no = next_row - 2  # Because row 3 is S.No. = 1

                # Write to Excel
                sheet.cell(row=next_row, column=1).value = serial_no
                sheet.cell(row=next_row, column=2).value = name
                sheet.cell(row=next_row, column=3).value = age
                sheet.cell(row=next_row, column=4).value = email
                sheet.cell(row=next_row, column=5).value = phone

                wb.save(file_path)

                Label(window, text="Registration Successful", font="ariel 12", fg="green").place(x=385, y=100)
                # clear()  # Clear fields after success
            else:
                messagebox.showerror("Error", "Excel file not found.")
        except Exception as e:
            messagebox.showerror("Error", f"Could not write to Excel file:\n{e}")
    
    # Write data to a file
    # with open (name +".txt", "w") as f:
    #     f.write( "Name: "+name+"\n")
    #     f.write("Age: "+age+ "\n")
    #     f.write("E-Mail: "+email+"\n")
    #     f.write( "Phone: "+phone+ "\n")

def clear():
    name_enter.delete (0,END)
    age_enter.delete (0, END) 
    email_enter.delete (0, END) 
    phone_enter.delete(0, END)
    
    Label(window, text="Fields are Cleared", font="ariel 10 bold", fg="red"). place (x=100, y=350)

def open_browser():
    wb.open_new("https://sites.google.com/site/ehsanafzal")
# Header
header = PhotoImage(file="rbyc.logo.png")
Label(window, image=header, bg="white").pack(fill="both")
# Field Lables
Label(window, text="Name", font="ariel 14").place(x=60,y=110)
Label(window, text="Age", font="ariel 14").place(x=60,y=150)
Label(window, text="E-mail", font="ariel 14").place(x=60,y=190)
Label(window, text="Phone", font="ariel 14").place(x=60, y=230)
# Creator Signature
Label(window, text="Copyright © Ehsan 2025", font="ariel 8") .place(x=414,y=350)
Label (window, text="All Rights Reserved : Ehsan Afzal", font="ariel 8").place(x=365, y=370)
# Putting logo
logo=PhotoImage (file="ea.logo.png")
Label(window, image=logo) .place (x=540, y=340)
# Entered item collected in String Variable

name_info=StringVar ()
age_info=StringVar ()
email_info=StringVar ()
phone_info=StringVar ()
# Entery Field created
name_enter=Entry(font="10", bd=2, textvariable=name_info)  # User input data will be stored in a dictionary 
name_enter.place(x=150, y=112) 
age_enter=Entry(font="10", bd=2, textvariable=age_info) # convert dictionary to data
age_enter.place(x=150, y=152)
email_enter=Entry (font="10", bd=2, textvariable=email_info) 
email_enter.place(x=150, y=192) 
phone_enter=Entry (font="10", bd=2, textvariable=phone_info) 
phone_enter.place(x=150, y=232)
# button
Button(window, text="Register", font="20", bd=2, bg="light green",command=register).place(x=260, y=280)
Button(window, text="Clear", font= "20", bd=2, bg="pink", command=clear).place(x=150,y=280)
Button(window, text="Reach out to the Developer", font="ariel 8", fg="blue", bd=2, command=open_browser).place(x=10,y=370)
# Closing mainloop
window.mainloop()